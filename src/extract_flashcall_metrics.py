"""
extract_flashcall_metrics.py
=================================
Extract all metric names from the Flashcall CRM Daily Report PDF and export
them as a formatted Excel workbook.

=== Data Source
- Raw:    data/raw/Flashcall_CRM_Daily_Report_20260810.pdf
- Output: data/processed/flashcall_crm_metrics.xlsx

=== Output Columns
序号 | 指标名称 | 中文翻译 | 类型(指标/分组标题)

=== Verification
Every metric name in the built-in list is looked up in the extracted PDF
text (whitespace-insensitive). Missing names trigger a warning so the
list can be kept in sync with future report versions.
"""

# === Imports
import os
import re

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# === Global Constants
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
PDF_PATH = os.path.join(PROJECT_ROOT, "data", "raw", "Flashcall_CRM_Daily_Report_20260810.pdf")
OUTPUT_PATH = os.path.join(PROJECT_ROOT, "data", "processed", "flashcall_crm_metrics.xlsx")

# (english name, chinese translation, kind) in report order
# kind: "metric" = data row, "group" = section header row
METRICS = [
    # Check-in (daily average)
    ("Check-in rep# (per day)", "每日打卡代表数", "metric"),
    ("Check-in rate (per day, excl leave etc.)", "每日打卡率（不含请假等）", "metric"),
    ("Average check-in time", "平均打卡时间", "metric"),
    ("Before 9 am Subtotal", "9点前打卡占比小计", "metric"),
    ("Before 7 am", "7点前打卡占比", "metric"),
    ("7-8 am", "7-8点打卡占比", "metric"),
    ("8-9 am", "8-9点打卡占比", "metric"),
    ("9-10 am", "9-10点打卡占比", "metric"),
    ("10-11 am", "10-11点打卡占比", "metric"),
    ("11 am -12 pm", "11-12点打卡占比", "metric"),
    ("After 12 pm", "12点后打卡占比", "metric"),
    # Rep Department meetings
    ("Reps by Dpt meeting frequency", "按部门会议频率划分的代表分布", "group"),
    ("Reps with >=2 (# / % of total)", "部门会议≥2次的代表（人数/占比）", "metric"),
    ("Reps with 0/1 (# / % of total)", "部门会议0或1次的代表（人数/占比）", "metric"),
    # Meeting & HCP Coverage
    ("# of meetings (per day)", "每日会议数", "metric"),
    ("# of meetings", "会议数", "metric"),
    ("Meeting by Product", "按产品划分的会议数分布", "group"),
    # Rep Coaching
    ("Reps by coaching frequency", "按辅导频率划分的代表分布", "group"),
    ("Reps with >=1 (# / % of total)", "接受辅导≥1次的代表（人数/占比）", "metric"),
    ("Reps with 0 (# / % of total)", "未接受辅导的代表（人数/占比）", "metric"),
    # HCP visit (not linked to incentives)
    ("# of reps inputting visit per day", "每日录入拜访的代表数", "metric"),
    ("% of total reps", "占代表总数比例", "metric"),
    ("# of HCP visited by reps per day", "代表每日拜访的HCP数", "metric"),
]

HEADER_FILL = "002577"
HEADER_FONT_COLOR = "FFFFFF"


def _load_pdf_text():
    """Extract all text from the source PDF.

    Returns:
        str: Concatenated text of all pages.
    """
    import PyPDF2

    if not os.path.exists(PDF_PATH):
        raise FileNotFoundError(f"PDF not found: {PDF_PATH}")
    reader = PyPDF2.PdfReader(PDF_PATH)
    pages = [page.extract_text() or "" for page in reader.pages]
    print(f"Loaded {os.path.basename(PDF_PATH)} | pages: {len(pages)}")
    return "\n".join(pages)


def _verify_metrics(text):
    """Check every metric name appears in the PDF text.

    Whitespace is removed from both sides so spacing artifacts from PDF
    text extraction do not break the lookup.

    Args:
        text: Full extracted PDF text.

    Returns:
        list[str]: Names that were NOT found (empty when all present).
    """
    normalized = re.sub(r"\s+", "", text)
    missing = []
    for name, _, _ in METRICS:
        if re.sub(r"\s+", "", name) not in normalized:
            missing.append(name)
    return missing


def _style_excel(writer, n_rows):
    """Apply Microsoft YaHei font, centered alignment and borders.

    Args:
        writer: Openpyxl ExcelWriter that already contains the sheet.
        n_rows: Number of data rows written (excluding header).
    """
    ws = writer.sheets["指标名称"]
    thin_side = Side(style="thin")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_font = Font(name="微软雅黑", bold=True, color=HEADER_FONT_COLOR)
    header_fill = PatternFill(
        start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid"
    )
    body_font = Font(name="微软雅黑", size=10)
    center = Alignment(horizontal="center", vertical="center")
    # Header row: bold white text on dark blue fill
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    # Data rows: centered, bordered
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        for cell in row:
            cell.font = body_font
            cell.alignment = center
            cell.border = border
    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12


def export_metrics():
    """Extract metric names from the PDF and write the Excel workbook."""
    text = _load_pdf_text()
    missing = _verify_metrics(text)
    if missing:
        print(f"Warning: {len(missing)} metric(s) NOT found in PDF:")
        for name in missing:
            print(f"  - {name}")
    else:
        print("Verification: all 23 metric names found in PDF")

    df_metrics = pd.DataFrame(
        [
            {"序号": i, "指标名称": name, "中文翻译": zh, "类型": kind}
            for i, (name, zh, kind) in enumerate(METRICS, start=1)
        ]
    )
    df_metrics.loc[df_metrics["类型"] == "group", "类型"] = "分组标题"
    df_metrics.loc[df_metrics["类型"] == "metric", "类型"] = "指标"

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        df_metrics.to_excel(writer, sheet_name="指标名称", index=False)
        _style_excel(writer, len(df_metrics))
    print("Generated flashcall_crm_metrics.xlsx")
    print("=" * 60)
    print(f"Metrics: {len(df_metrics)} rows | {OUTPUT_PATH}")


if __name__ == "__main__":
    export_metrics()
